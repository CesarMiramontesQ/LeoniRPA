"""Pruebas unitarias de exportación BOM a Excel (sin BD)."""
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.bom.export_excel import (
    BOM_EXPORT_COLUMNS,
    build_bom_export_xlsx,
    validate_row_counts,
    _format_plant,
    _to_excel_number,
)


def test_format_plant_matches_reference():
    assert _format_plant("US10", "1", "01") == "US10 / 1 / 01"
    assert _format_plant("US10", None, "01") == "US10 / 01"
    assert _format_plant(None, None, None) == ""


def test_to_excel_number_preserves_numeric_types():
    assert _to_excel_number(Decimal("9.14")) == 9.14
    assert _to_excel_number(1000) == 1000.0
    assert _to_excel_number(None) is None


def test_validate_row_counts_matches_components_per_part():
    filas = [
        {"Parte No": "520000001", "Material": "A"},
        {"Parte No": "520000001", "Material": "B"},
        {"Parte No": "520000003", "Material": "C"},
    ]
    assert validate_row_counts(filas, {"520000001": 2, "520000003": 1}) is True
    assert validate_row_counts(filas, {"520000001": 8, "520000003": 1}) is False


def test_build_bom_export_xlsx_one_row_per_component_no_merged_cells():
    filas = [
        {
            "Parte No": "520000001",
            "Plant": "US10 / 1 / 01",
            "Description": "LEONI Adascar Sensor 1925 2x2,5 + WSS",
            "Base Mts": 1000.0,
            "Req D": 1000.0,
            "Material": "310004003",
            "Description Material": "CU - GIESSWALZDR BL 8.00MM V1.2",
            "Qty": 9.14,
            "Measure": "KG",
        },
        {
            "Parte No": "520000001",
            "Plant": "US10 / 1 / 01",
            "Description": "LEONI Adascar Sensor 1925 2x2,5 + WSS",
            "Base Mts": 1000.0,
            "Req D": 1000.0,
            "Material": "310905000",
            "Description Material": "ZINN BLOECKE",
            "Qty": 0.12,
            "Measure": "KG",
        },
        {
            "Parte No": "520000003",
            "Plant": "US10 / 1 / 01",
            "Description": "Otra parte",
            "Base Mts": 1000.0,
            "Req D": 1000.0,
            "Material": "340060066",
            "Description Material": "2X-112-HEFR",
            "Qty": 3.88,
            "Measure": "KG",
        },
    ]

    data = build_bom_export_xlsx(filas)
    wb = load_workbook(BytesIO(data))
    ws = wb.active

    assert [cell.value for cell in ws[1]] == BOM_EXPORT_COLUMNS
    assert ws.max_row == 1 + len(filas)
    assert ws.merged_cells.ranges == set() or len(ws.merged_cells.ranges) == 0

    # Parte 520000001 aparece tantas veces como componentes (2)
    parte_col = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert parte_col.count("520000001") == 2
    assert parte_col.count("520000003") == 1

    # Qty numérico
    assert ws.cell(row=2, column=8).value == 9.14
    assert isinstance(ws.cell(row=2, column=8).value, float)
    assert ws.cell(row=2, column=6).value == "310004003"
    assert ws.cell(row=2, column=2).value == "US10 / 1 / 01"
