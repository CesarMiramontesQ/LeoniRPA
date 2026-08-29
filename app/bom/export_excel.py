"""Exportación Excel del BOM plano (una fila por componente)."""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.db.models import Bom, BomItem, BomRevision, Parte

# Columnas en el orden de la plantilla de referencia SAP/BOM.
BOM_EXPORT_COLUMNS = [
    "Parte No",
    "Plant",
    "Description",
    "Base Mts",
    "Req D",
    "Material",
    "Description Material",
    "Qty",
    "Measure",
]

# Encabezado azul oscuro similar a la referencia.
_HEADER_FILL = PatternFill(fill_type="solid", start_color="003D89", end_color="003D89")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _format_plant(plant: Optional[str], usage: Optional[str], alternative: Optional[str]) -> str:
    """Formato Plant de la referencia: 'US10 / 1 / 01'."""
    parts = [str(p).strip() for p in (plant, usage, alternative) if p is not None and str(p).strip()]
    return " / ".join(parts)


def _to_excel_number(value: Any) -> Optional[float]:
    """Convierte Decimal/int/float a float para Excel; None si no hay valor."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


async def fetch_bom_export_rows(db: AsyncSession) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Obtiene, en una sola consulta, todas las filas del BOM vigente
    (revisión con effective_to IS NULL) para números de parte finales (padre del BOM).

    Retorna (filas, resumen) donde cada fila mapea las columnas del Excel y el resumen
    incluye conteos para validar filas vs componentes y partes sin ítems.
    """
    padre = aliased(Parte)
    componente = aliased(Parte)

    query = (
        select(
            padre.numero_parte.label("parte_no"),
            Bom.plant,
            Bom.usage,
            Bom.alternative,
            padre.descripcion.label("description"),
            Bom.base_qty,
            Bom.reqd_qty,
            componente.numero_parte.label("material"),
            componente.descripcion.label("description_material"),
            BomItem.qty,
            BomItem.measure,
            BomItem.item_no,
        )
        .select_from(BomItem)
        .join(BomRevision, BomRevision.id == BomItem.bom_revision_id)
        .join(Bom, Bom.id == BomRevision.bom_id)
        .join(padre, padre.id == Bom.parte_id)
        .join(componente, componente.id == BomItem.componente_id)
        .where(BomRevision.effective_to.is_(None))
        .order_by(
            padre.numero_parte,
            Bom.plant,
            Bom.usage,
            Bom.alternative,
            BomItem.item_no.nulls_last(),
            BomItem.id,
            componente.numero_parte,
        )
    )
    result = await db.execute(query)
    raw_rows = result.all()

    filas: List[Dict[str, Any]] = []
    componentes_por_parte: Dict[str, int] = {}

    for row in raw_rows:
        parte_no = (row.parte_no or "").strip()
        filas.append(
            {
                "Parte No": parte_no,
                "Plant": _format_plant(row.plant, row.usage, row.alternative),
                "Description": (row.description or "").strip() if row.description else "",
                "Base Mts": _to_excel_number(row.base_qty),
                "Req D": _to_excel_number(row.reqd_qty),
                "Material": (row.material or "").strip() if row.material else "",
                "Description Material": (
                    (row.description_material or "").strip() if row.description_material else ""
                ),
                "Qty": _to_excel_number(row.qty),
                "Measure": (row.measure or "").strip() if row.measure else "",
            }
        )
        if parte_no:
            componentes_por_parte[parte_no] = componentes_por_parte.get(parte_no, 0) + 1

    # Partes con BOM/revisión vigente pero sin ítems (productos sin componentes asociados).
    revision_con_items = (
        select(BomItem.id)
        .where(BomItem.bom_revision_id == BomRevision.id)
        .correlate(BomRevision)
        .exists()
    )
    sin_items_q = (
        select(func.count(func.distinct(Bom.parte_id)))
        .select_from(Bom)
        .join(BomRevision, BomRevision.bom_id == Bom.id)
        .where(
            BomRevision.effective_to.is_(None),
            ~revision_con_items,
        )
    )
    partes_sin_items = int((await db.execute(sin_items_q)).scalar() or 0)

    # Partes finales válidas sin ningún BOM (para mensaje de UI).
    sin_bom_subq = select(Bom.id).where(Bom.parte_id == Parte.id).exists()
    sin_bom_q = await db.execute(
        select(func.count(Parte.id)).where(Parte.valido.is_(True), ~sin_bom_subq)
    )
    partes_sin_bom = int(sin_bom_q.scalar() or 0)

    resumen = {
        "total_filas": len(filas),
        "total_partes_con_componentes": len(componentes_por_parte),
        "componentes_por_parte": componentes_por_parte,
        "partes_sin_items_revision_vigente": partes_sin_items,
        "partes_validas_sin_bom": partes_sin_bom,
    }
    return filas, resumen


def build_bom_export_xlsx(filas: List[Dict[str, Any]]) -> bytes:
    """Genera un .xlsx con una fila por componente; sin celdas combinadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    for col_idx, header in enumerate(BOM_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    numeric_cols = {"Base Mts", "Req D", "Qty"}

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, header in enumerate(BOM_EXPORT_COLUMNS, start=1):
            value = fila.get(header)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if header in numeric_cols and value is not None:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Anchos aproximados según la referencia visual.
    widths = {
        "A": 14,  # Parte No
        "B": 16,  # Plant
        "C": 42,  # Description
        "D": 12,  # Base Mts
        "E": 10,  # Req D
        "F": 14,  # Material
        "G": 36,  # Description Material
        "H": 12,  # Qty
        "I": 10,  # Measure
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def validate_row_counts(
    filas: List[Dict[str, Any]], componentes_por_parte: Dict[str, int]
) -> bool:
    """Verifica que las filas generadas por parte coincidan con el conteo de componentes."""
    counts: Dict[str, int] = {}
    for fila in filas:
        pn = fila.get("Parte No") or ""
        if pn:
            counts[pn] = counts.get(pn, 0) + 1
    return counts == componentes_por_parte
